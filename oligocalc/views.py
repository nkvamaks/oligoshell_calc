import io
import numpy as np
from django.shortcuts import render, redirect
from django.core.mail import EmailMessage
from django.conf import settings
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse, reverse_lazy
from django.views.generic import FormView, DeleteView
from django.template.loader import render_to_string
from meta.views import MetadataMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from allauth.account.models import EmailAddress
from openpyxl import Workbook


from . import forms
from . import utils
from . import tm
from . import taqman_find_utils
from . import metadata


class CalcView(MetadataMixin, FormView):
    template_name = 'oligocalc/calculator.html'
    form_class = forms.CalcForm
    success_url = reverse_lazy('oligocalc:calculator')
    title = 'Oligonucleotide Properties Calculator'
    description = 'Calculate properties of nucleic acids online. Determine melting temperature, molecular weight, extinction coefficient for DNA, RNA, PMO and therapeutic oligonucleotides.'
    keywords = ['oligonucleotide properties calculator', 'oligo calculator', 'oligonucleotide calculator', 'oligocalculator', 'oligo mass calculator', 'oligocalc',
                'oligonucleotide properties', 'nucleic acids', 'melting temperature', 'Tm', 'extinction coefficient',
                'exact mass', 'molecular weight', 'DNA', 'RNA', 'PMO', 'morpholino', 'modified oligonucleotides', 'therapeutic oligonucleotides',
                'minor groove binder', 'MGB', 'modifications', 'conjugates', 'bioconjugates']

    def form_valid(self, form):
        if 'export_excel_neg' in self.request.POST:
            sequence_data = self.calculate_results(form)
            ms_frag_array = sequence_data.get('mass_fragments_array_neg')
            seq_wo_phosph_tup = sequence_data.get('seq_wo_phosph_tup')
            return self.export_to_excel(ms_frag_array, seq_wo_phosph_tup, mode='neg')
        if 'export_excel_pos' in self.request.POST:
            sequence_data = self.calculate_results(form)
            ms_frag_array = sequence_data.get('mass_fragments_array_pos')
            seq_wo_phosph_tup = sequence_data.get('seq_wo_phosph_tup')
            return self.export_to_excel(ms_frag_array, seq_wo_phosph_tup, mode='pos')

        context = self.get_context_data(form=form)
        context.update(self.calculate_results(form))
        return self.render_to_response(context)

    def export_to_excel(self, ms_frag_array, seq_wo_phosph_tup, mode):
        """
        Create an Excel file (using openpyxl) and return it as an attachment.
        """
        wb = Workbook()
        wb.remove(wb.active)

        # If ms_frag_array is empty or None, handle the "no data" scenario
        if ms_frag_array is None or ms_frag_array.size == 0:
            ws_empty = wb.create_sheet(title='NoData')
            ws_empty.append(['No MS2 data'])
        else:
            for idx, charge_state_data in enumerate(ms_frag_array, start=1):
                if mode == 'neg':
                    ws = wb.create_sheet(title=f'z-{idx}')
                else:
                    ws = wb.create_sheet(title=f'z+{idx}')
                headers = ['d', 'd-NMe2', 'c', 'b', 'a', 'a-B', "5'-Index", 'Sequence', "3'-Index", 'w', 'x', 'y', 'z', 'z-NMe2']
                ws.append(headers)

                n = len(charge_state_data)

                for row_idx, row_data_tup in enumerate(charge_state_data):
                    d, c, b, a, a_B, w, x, y, z, d_backbone, z_backbone = row_data_tup

                    row_data = [d, d_backbone, c, b, a, a_B,
                                row_idx + 1,                 # 5'-Index
                                seq_wo_phosph_tup[row_idx],  # Sequence
                                n - row_idx,                 # 3'-Index
                                w, x, y, z, z_backbone]
                    ws.append(row_data)

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = 'ms2_frag_neg.xlsx' if mode == 'neg' else 'ms2_frag_pos.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def calculate_results(self, form):
        sequence, mass_monoisotopic, concentration_molar, concentration_mass = '', 0, 0, 0
        quantity, melting_t, esi_series_neg, esi_series_pos = 0, -1, [], []
        mass_fragments_array_neg, mass_fragments_array_pos = None, None

        btnradio = form.cleaned_data['btnradio']
        if btnradio == 'dna':
            sequence = utils.dna2mix(form.cleaned_data['sequence'])
        elif btnradio == 'rna':
            sequence = utils.rna2mix(form.cleaned_data['sequence'])
        else:
            sequence = form.cleaned_data['sequence']

        sequence_dna = utils.mix2dna_wo_mod_phosph(utils.wo_phosph(sequence))
        sequence_dna_rev_compl = utils.rev_compl(sequence_dna)
        seq_wo_phosph_tup = utils.sequence_split(utils.wo_phosph(sequence))
        length = utils.get_length(sequence)
        volume = form.cleaned_data['volume']
        absorbance260 = form.cleaned_data['absorbance260']
        mv_conc = form.cleaned_data['mv_conc']
        dv_conc = form.cleaned_data['dv_conc']
        dntp_conc = form.cleaned_data['dntp_conc']
        dna_conc = form.cleaned_data['dna_conc']
        target = form.cleaned_data['target']

        epsilon260 = utils.get_extinction(sequence)
        gc_content = utils.gc_content(seq_wo_phosph_tup)
        brutto_formula = utils.get_formula(sequence)
        mass_average = utils.get_mass_avg(sequence)

        nmol_OD260, ug_OD260 = utils.calculate_od260(epsilon260, mass_average)

        if not utils.contain_degenerate_nucleotide(sequence):
            mass_monoisotopic = utils.get_mass_monoisotopic(sequence)

        esi_series_neg, esi_series_pos = utils.calculate_esi_series(sequence, length, mass_average, mass_monoisotopic)

        if absorbance260 and epsilon260:
            concentration_molar = absorbance260 / epsilon260 * 1e6
            concentration_mass = concentration_molar * mass_average / 1000

        if concentration_molar and volume:
            quantity = concentration_molar * volume

        if not utils.contain_degenerate_nucleotide(sequence):
            mass_fragments_array_neg, mass_fragments_array_pos = utils.calculate_mass_fragments(sequence)

        melting_t = tm.calculate_melting_temp(seq_wo_phosph_tup, sequence, target, dna_conc, mv_conc, dv_conc,
                                              dntp_conc)

        return {
            'form': form,
            'mv_conc': mv_conc,
            'dv_conc': dv_conc,
            'dntp_conc': dntp_conc,
            'dna_conc': dna_conc,
            'sequence': sequence,
            'seq_wo_phosph_tup': seq_wo_phosph_tup,
            'sequence_dna': sequence_dna,
            'sequence_dna_rev_compl': sequence_dna_rev_compl,
            'volume': volume,
            'epsilon260': epsilon260,
            'absorbance260': absorbance260,
            'length': length,
            'charge': range(1, length),
            'concentration_molar': concentration_molar,
            'concentration_mass': concentration_mass,
            'quantity': quantity,
            'mass_monoisotopic': mass_monoisotopic,
            'mass_average': mass_average,
            'esi_series_neg': esi_series_neg,
            'esi_series_pos': esi_series_pos,
            'mass_fragments_array_neg': mass_fragments_array_neg,
            'mass_fragments_array_pos': mass_fragments_array_pos,
            'melting_t': melting_t,
            'gc_content': int(gc_content * 100),
            'brutto_formula': brutto_formula,
            'nmol_OD260': nmol_OD260,
            'ug_OD260': ug_OD260,
            'contain_degenerate_nucleotide': utils.contain_degenerate_nucleotide(sequence)}


def about(request):
    with open(settings.BASE_DIR / 'README.md', 'r') as fh:
        return render(request, 'oligocalc/about.html', {'about_osh_calc': ''.join(line for line in fh),
                                                        'meta': metadata.get_about_meta(request)})


def contact(request):
    if request.method == "POST":
        form = forms.ContactForm(request.POST)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']
            name = form.cleaned_data['name']
            reply_to = form.cleaned_data['reply_to']

            body = 'From:\t{}\nEmail:\t{}\nMessage:\t\t\n\n{}\n'.format(name, reply_to, message)

            email = EmailMessage(
                subject=subject,
                body=body,
                from_email='OligoShell Contact Form',
                to=[settings.EMAIL_HOST_USER, ],
                reply_to=[reply_to, ],
            )
            email.send()
            return HttpResponseRedirect(reverse('oligocalc:success'))

    else:
        form = forms.ContactForm()
    return render(request, 'oligocalc/contact.html', {'form': form,
                                                      'meta': metadata.get_contact_meta(request)})


def success(request):
    return render(request, 'oligocalc/successfully_sent.html')


def modifications(request):
    return render(request, 'oligocalc/modifications.html', {'meta': metadata.get_modifications_meta(request)})


class TaqManFindView(MetadataMixin, FormView):
    template_name = 'oligocalc/taqman_find.html'
    form_class = forms.TaqManFindForm
    title = 'TaqMan Assay Finder'
    description = 'Find primers and probes for TaqMan assays within your target sequence by providing masses of primers and probe, chemical modifications, length of the amplicon, and reference sequence (RefSeq).'
    keywords = ['TaqMan probe', 'primer', 'real-time PCR', 'qPCR', 'PCR', 'RT-qPCR', 'probe finder', 'PCR probes',
                'minor groove binder', 'MGB', 'modifications', 'exact mass', 'TaqMan assay', 'gene expression',
                'gene expression analysis', 'transcript', 'oligonucleotide', 'DNA', 'RNA']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_taqman'] = self.get_form(self.form_class)
        return context

    def post(self, request, *args, **kwargs):
        form_taqman = self.get_form(self.form_class)
        fasta = ''

        if form_taqman.is_valid():
            fasta = form_taqman.cleaned_data['fasta']
            fasta_seq = taqman_find_utils.simple_fasta_parser(fasta)
            primer1 = form_taqman.cleaned_data['primer1']
            primer2 = form_taqman.cleaned_data['primer2']
            probe = form_taqman.cleaned_data['probe']
            probe_dye = form_taqman.cleaned_data['probe_dye']
            amp_size = form_taqman.cleaned_data['amp_size']
            fasta_seq_len = len(fasta_seq)
            fasta_seq_rev_compl = utils.rev_compl(fasta_seq)

            composition_list_primer1 = taqman_find_utils.brutforce_oligo_composition(primer1)
            composition_list_primer2 = taqman_find_utils.brutforce_oligo_composition(primer2)
            composition_list_probe = taqman_find_utils.brutforce_oligo_composition(
                probe - taqman_find_utils.mass_probe_dye[probe_dye])

            seq_list_F_set1 = taqman_find_utils.find_seq_in_fasta_from_list(composition_list_primer1, fasta_seq)
            seq_list_F_set2 = taqman_find_utils.find_seq_in_fasta_from_list(composition_list_primer2, fasta_seq)
            seq_list_R_set1 = taqman_find_utils.find_seq_in_fasta_from_list(composition_list_primer2,
                                                                            fasta_seq_rev_compl)
            seq_list_R_set2 = taqman_find_utils.find_seq_in_fasta_from_list(composition_list_primer1,
                                                                            fasta_seq_rev_compl)

            match_set1 = taqman_find_utils.find_matching_pair(seq_list_F_set1, seq_list_R_set1, amp_size, fasta_seq_len)
            match_set2 = taqman_find_utils.find_matching_pair(seq_list_F_set2, seq_list_R_set2, amp_size, fasta_seq_len)

            results = []
            for seqF_seq, seqF_pos, seqR_seq, seqR_pos in match_set1 + match_set2:
                amplicon = fasta_seq[seqF_pos:seqF_pos + amp_size]
                amplicon_rev_compl = utils.rev_compl(amplicon)

                seqF_mass = utils.get_mass_monoisotopic(utils.dna2mix(seqF_seq))
                seqR_mass = utils.get_mass_monoisotopic(utils.dna2mix(seqR_seq))

                seqF_melting_t = tm.calc_tm(seq=utils.dna2mix(seqF_seq),
                                            target='dna',
                                            dna_conc=0.2,
                                            mv_conc=50,
                                            dv_conc=3,
                                            dntp_conc=0.8
                                            )
                seqR_melting_t = tm.calc_tm(seq=utils.dna2mix(seqR_seq),
                                            target='dna',
                                            dna_conc=0.2,
                                            mv_conc=50,
                                            dv_conc=3,
                                            dntp_conc=0.8
                                            )

                seq_list_Pf = taqman_find_utils.find_seq_in_fasta_from_list(composition_list_probe, amplicon)
                seq_list_Pr = taqman_find_utils.find_seq_in_fasta_from_list(composition_list_probe, amplicon_rev_compl)

                seq_list_Pf_mass_tm = taqman_find_utils.calculate_mass_tm(seq_list_Pf, probe_dye)
                seq_list_Pr_mass_tm = taqman_find_utils.calculate_mass_tm(seq_list_Pr, probe_dye)

                master_list_Pf = [x + y for x, y in zip(seq_list_Pf, seq_list_Pf_mass_tm)]
                master_list_Pr = [x + y for x, y in zip(seq_list_Pr, seq_list_Pr_mass_tm)]

                assay = taqman_find_utils.get_taqman_assay(seqF_seq, seqR_seq, amplicon, seq_list_Pf, seq_list_Pr,
                                                           seqF_pos)

                results.append((seqF_seq, seqF_pos, seqF_mass, seqF_melting_t,
                                seqR_seq, seqR_pos, seqR_mass, seqR_melting_t,
                                amplicon, amplicon_rev_compl,
                                master_list_Pf, master_list_Pr,
                                assay))
            if not results:
                results = -1
            context = self.get_context_data()
            context.update({
                'form_taqman': form_taqman,
                'fasta': fasta,
                'results': results,
                'amp_size': amp_size,
            })
            return self.render_to_response(context)

        context = self.get_context_data()
        context.update({
            'form_taqman': form_taqman,
            'fasta': fasta})
        return self.render_to_response(context)


def robots_txt(request):
    sitemap_url = request.build_absolute_uri(reverse('oligocalc:django.contrib.sitemaps.views.sitemap'))
    content = render_to_string('oligocalc/robots.txt', {'sitemap_url': sitemap_url})
    return HttpResponse(content, content_type='text/plain')


class ProfileDetails(MetadataMixin, LoginRequiredMixin, DeleteView):
    model = User
    template_name = 'oligocalc/profile_details.html'
    success_url = reverse_lazy('oligocalc:calculator')
    title = 'Profile Details'
    description = 'Manage your personal information, associated email addresses, and connected social accounts securely. Update your profile details or delete your account with ease.'
    keywords = ['user profile', 'account management', 'associated emails', 'social accounts', 'manage profile', 'personal information', 'update profile',
                'oligonucleotide', 'DNA', 'RNA', 'PMO']

    def get_object(self, **kwargs):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['email_addresses'] = EmailAddress.objects.filter(user=self.request.user)
        context['delete_account_form'] = forms.DeleteAccountForm()
        return context

    def post(self, request, *args, **kwargs):
        form = forms.DeleteAccountForm(request.POST)
        if form.is_valid():
            user = self.get_object()
            user.delete()
            return redirect(reverse_lazy('account_logout'))
        else:
            return self.get(request, *args, **kwargs)

